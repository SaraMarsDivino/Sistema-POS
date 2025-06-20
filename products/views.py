#views.py/products
from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from django.db.models import Q  # Para filtros de búsqueda
from .models import Product
from .forms import ProductForm
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from .models import Product

# products/views.py
def product_management(request):
    query = request.GET.get('search', '')
    products = Product.objects.filter(Q(nombre__icontains=query) | Q(producto_id__icontains=query))

    paginator = Paginator(products, 10)  # Paginación
    page = request.GET.get('page')
    products_page = paginator.get_page(page)

    return render(request, 'products/product_management.html', {
        'products': products_page,
        'search_query': query
    })

from django.contrib import messages

def upload_products(request):
    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, 'No se subió ningún archivo.')
            return redirect('upload_products')

        file = request.FILES['file']
        if not file.name.endswith('.xlsx'):
            messages.error(request, 'Por favor sube un archivo válido en formato Excel (.xlsx).')
            return redirect('upload_products')

        try:
            workbook = load_workbook(file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[1]:  # Validar que haya datos en el nombre
                    producto_id = row[0]
                    nombre = row[1]
                    descripcion = row[2]
                    cantidad = row[3] or 0
                    precio_compra = row[4] or 0.00
                    precio_venta = row[5] or 0.00
                    codigo_barras = row[6]

                    # Generar producto_id si no está definido
                    if not producto_id:
                        producto_id = f"PRD-{Product.objects.count() + 1}"

                    # Asegurar unicidad del producto_id
                    original_id = producto_id
                    counter = 2
                    while Product.objects.filter(producto_id=producto_id).exists():
                        producto_id = f"{original_id} ({counter})"
                        counter += 1

                    Product.objects.update_or_create(
                        producto_id=producto_id,
                        defaults={
                            'nombre': nombre,
                            'descripcion': descripcion,
                            'cantidad': cantidad,
                            'precio_compra': precio_compra,
                            'precio_venta': precio_venta,
                            'codigo_barras': codigo_barras,
                        }
                    )
            
            # Agregar mensaje de éxito
            messages.success(request, 'La plantilla se subió y procesó con éxito.')
            return redirect('product_management')
        except Exception as e:
            messages.error(request, f'Error procesando el archivo: {str(e)}')
            return redirect('upload_products')

    return render(request, 'products/upload_products.html')



def download_template(request):
    # Crear un archivo Excel como plantilla
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_productos.xlsx"'

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Productos'
    headers = ['Producto ID', 'Nombre', 'Descripción', 'Cantidad', 'Precio de Compra', 'Precio de Venta', 'Código de Barras']
    sheet.append(headers)

    workbook.save(response)
    return response



# products/views.py
def create_or_edit_product(request, product_id=None):
    product = get_object_or_404(Product, id=product_id) if product_id else None
    form = ProductForm(request.POST or None, instance=product)
    title = 'Editar Producto' if product_id else 'Crear Producto'

    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Los cambios se guardaron con éxito.')
        if 'save_and_list' in request.POST:
            return redirect('product_management')
        return redirect('edit_product', product.id if product else form.instance.id)

    return render(request, 'products/product_form.html', {
        'form': form, 'title': title, 'product': product
    })





def delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Producto eliminado exitosamente.')
        return redirect('product_management')
    return render(request, 'products/delete_product.html', {'product': product})
